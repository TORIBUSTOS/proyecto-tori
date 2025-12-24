"""
Endpoints de exportación (ETAPA 7)
- PDF: Reporte ejecutivo
- Excel: Movimientos filtrados
- Excel Ejecutivo: 5 hojas completas (ETAPA 7.B)
"""

from fastapi import HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import date, datetime
from io import BytesIO
import pandas as pd
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from backend.database.connection import get_db
from backend.models.movimiento import Movimiento
from backend.core.reportes import generar_reporte_ejecutivo


def generar_pdf_reporte(reporte: dict) -> BytesIO:
    """
    Genera un PDF del reporte ejecutivo usando ReportLab
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)

    # Contenedor de elementos
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#764ba2'),
        spaceAfter=12,
        spaceBefore=20
    )

    # Título
    elements.append(Paragraph("TORO Investment Manager", title_style))
    elements.append(Paragraph(f"Reporte Ejecutivo - {reporte['periodo']}", styles['Heading2']))
    elements.append(Spacer(1, 0.3*inch))

    # KPIs
    elements.append(Paragraph("Resumen del Período", heading_style))
    kpis = reporte['kpis']
    kpis_data = [
        ['KPI', 'Valor'],
        ['Ingresos Totales', f"${kpis['ingresos_total']:,.2f}"],
        ['Egresos Totales', f"${kpis['egresos_total']:,.2f}"],
        ['Saldo Neto', f"${kpis['saldo_neto']:,.2f}"],
        ['Cantidad de Movimientos', str(kpis['cantidad_movimientos'])]
    ]

    kpis_table = Table(kpis_data, colWidths=[3*inch, 3*inch])
    kpis_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(kpis_table)
    elements.append(Spacer(1, 0.3*inch))

    # Saldos Bancarios
    elements.append(Paragraph("Saldos Bancarios", heading_style))
    saldos = reporte['saldos']
    saldos_data = [
        ['Concepto', 'Valor'],
        ['Saldo Inicial', f"${saldos['saldo_inicial']:,.2f}"],
        ['Ingresos del Período', f"${saldos['ingresos_total']:,.2f}"],
        ['Egresos del Período', f"${saldos['egresos_total']:,.2f}"],
        ['Variación', f"${saldos['variacion']:,.2f}"],
        ['Saldo Final', f"${saldos['saldo_final']:,.2f}"]
    ]

    saldos_table = Table(saldos_data, colWidths=[3*inch, 3*inch])
    saldos_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(saldos_table)
    elements.append(Spacer(1, 0.3*inch))

    # Clasificación
    elements.append(Paragraph("Clasificación de Movimientos", heading_style))
    clas = reporte['clasificacion']
    clas_data = [
        ['Concepto', 'Cantidad'],
        ['Total de Movimientos', str(clas['total_movimientos'])],
        ['Movimientos Clasificados', str(clas['clasificados'])],
        ['Sin Clasificar', str(clas['sin_clasificar'])],
        ['Porcentaje Clasificado', f"{clas['pct_clasificados']}%"]
    ]

    clas_table = Table(clas_data, colWidths=[3*inch, 3*inch])
    clas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    elements.append(clas_table)

    # Nueva página para desglose
    elements.append(PageBreak())

    # Desglose de Ingresos
    if reporte['desglose_ingresos']:
        elements.append(Paragraph("Desglose de Ingresos por Categoría", heading_style))
        ing_data = [['Categoría', 'Monto']]
        for item in reporte['desglose_ingresos']:
            ing_data.append([item['categoria'], f"${item['monto']:,.2f}"])

        ing_table = Table(ing_data, colWidths=[3*inch, 3*inch])
        ing_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(ing_table)
        elements.append(Spacer(1, 0.3*inch))

    # Desglose de Egresos
    if reporte['desglose_egresos']:
        elements.append(Paragraph("Desglose de Egresos por Categoría", heading_style))
        egr_data = [['Categoría', 'Monto']]
        for item in reporte['desglose_egresos']:
            egr_data.append([item['categoria'], f"${item['monto']:,.2f}"])

        egr_table = Table(egr_data, colWidths=[3*inch, 3*inch])
        egr_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightcoral),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(egr_table)

    # Generar PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


async def exportar_reporte_pdf(
    mes: Optional[str] = Query(None, description="Mes en formato YYYY-MM"),
    db: Session = Depends(get_db)
):
    """
    Endpoint: GET /api/reportes/pdf
    Exporta el reporte ejecutivo en formato PDF
    """
    try:
        # Generar reporte
        reporte = generar_reporte_ejecutivo(db, mes)

        # Generar PDF
        pdf_buffer = generar_pdf_reporte(reporte)

        # Nombre del archivo
        filename = f"reporte_{reporte['periodo'].replace(' ', '_')}.pdf"

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generando PDF: {str(e)}"
        )


async def exportar_movimientos_excel(
    fecha_desde: Optional[date] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[date] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    categoria: Optional[str] = Query(None, description="Filtrar por categoría"),
    mes: Optional[str] = Query(None, description="Filtrar por mes (YYYY-MM)"),
    db: Session = Depends(get_db)
):
    """
    Endpoint: GET /api/movimientos/excel
    Exporta movimientos filtrados a Excel
    """
    try:
        # Query base
        query = db.query(Movimiento)

        # Aplicar filtros
        if mes:
            # Parsear mes
            year, month = mes.split('-')
            query = query.filter(
                func.extract('year', Movimiento.fecha) == int(year),
                func.extract('month', Movimiento.fecha) == int(month)
            )
        else:
            if fecha_desde:
                query = query.filter(Movimiento.fecha >= fecha_desde)
            if fecha_hasta:
                query = query.filter(Movimiento.fecha <= fecha_hasta)

        if categoria:
            query = query.filter(Movimiento.categoria == categoria)

        # Ejecutar query
        movimientos = query.order_by(Movimiento.fecha.desc()).all()

        if not movimientos:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron movimientos con los filtros especificados"
            )

        # Crear DataFrame
        data = []
        for m in movimientos:
            data.append({
                'Fecha': m.fecha.strftime('%Y-%m-%d'),
                'Descripción': m.descripcion,
                'Monto': m.monto,
                'Saldo': m.saldo if m.saldo else '',
                'Categoría': m.categoria or 'SIN_CATEGORIA',
                'Subcategoría': m.subcategoria or '',
                'Confianza (%)': m.confianza_porcentaje or 0,
                'Persona/Empresa': m.persona_nombre or '',
                'Documento': m.documento or '',
                'Es DEBIN': 'Sí' if m.es_debin else 'No',
                'DEBIN ID': m.debin_id or ''
            })

        df = pd.DataFrame(data)

        # Generar Excel
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Movimientos')

            # Ajustar ancho de columnas
            worksheet = writer.sheets['Movimientos']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length + 2, 50)

        excel_buffer.seek(0)

        # Nombre del archivo
        if mes:
            filename = f"movimientos_{mes}.xlsx"
        elif fecha_desde and fecha_hasta:
            filename = f"movimientos_{fecha_desde}_{fecha_hasta}.xlsx"
        else:
            filename = "movimientos.xlsx"

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generando Excel: {str(e)}"
        )


def generar_excel_ejecutivo(reporte: dict, mes: str, db: Session) -> BytesIO:
    """
    Genera Excel Ejecutivo con 5 hojas (ETAPA 7.B)
    Hojas: Resumen, Ingresos, Egresos, Top Egresos, Sin Clasificar
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto

    # Fuente para encabezados
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # Parsear mes
    year, month = mes.split('-')

    # ============================================
    # HOJA 1: RESUMEN
    # ============================================
    ws_resumen = wb.create_sheet("Resumen")

    row = 1

    # Título
    ws_resumen['A1'] = "REPORTE EJECUTIVO"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['B1'] = reporte['periodo']
    row = 3

    # SALDOS BANCARIOS
    ws_resumen[f'A{row}'] = "SALDOS BANCARIOS"
    ws_resumen[f'A{row}'].font = header_font
    row += 1

    saldos = reporte['saldos']
    saldos_data = [
        ['Saldo Inicial', saldos['saldo_inicial']],
        ['Total Ingresos', saldos['ingresos_total']],
        ['Total Egresos', saldos['egresos_total']],
        ['Saldo Final', saldos['saldo_final']],
        ['Variacion del Mes', saldos['variacion']]
    ]

    for label, valor in saldos_data:
        ws_resumen[f'A{row}'] = label
        ws_resumen[f'B{row}'] = valor
        ws_resumen[f'B{row}'].number_format = '#,##0.00'
        row += 1

    row += 1

    # CLASIFICACION
    ws_resumen[f'A{row}'] = "CLASIFICACION"
    ws_resumen[f'A{row}'].font = header_font
    row += 1

    clas = reporte['clasificacion']
    clas_data = [
        ['Total Movimientos', clas['total_movimientos']],
        ['Clasificados', clas['clasificados']],
        ['Sin Clasificar', clas['sin_clasificar']],
        ['% Clasificados', f"{clas['pct_clasificados']}%"]
    ]

    for label, valor in clas_data:
        ws_resumen[f'A{row}'] = label
        ws_resumen[f'B{row}'] = valor
        row += 1

    row += 1

    # DESGLOSE INGRESOS
    ws_resumen[f'A{row}'] = "DESGLOSE INGRESOS"
    ws_resumen[f'A{row}'].font = header_font
    row += 1

    ws_resumen[f'A{row}'] = "Categoria/Subcategoria"
    ws_resumen[f'A{row}'].font = header_font
    ws_resumen[f'B{row}'] = "Monto"
    ws_resumen[f'B{row}'].font = header_font
    row += 1

    for item in reporte['desglose_ingresos']:
        ws_resumen[f'A{row}'] = item['categoria']
        ws_resumen[f'B{row}'] = item['monto']
        ws_resumen[f'B{row}'].number_format = '#,##0.00'
        row += 1

    row += 1

    # DESGLOSE EGRESOS
    ws_resumen[f'A{row}'] = "DESGLOSE EGRESOS"
    ws_resumen[f'A{row}'].font = header_font
    row += 1

    ws_resumen[f'A{row}'] = "Categoria/Subcategoria"
    ws_resumen[f'A{row}'].font = header_font
    ws_resumen[f'B{row}'] = "Monto"
    ws_resumen[f'B{row}'].font = header_font
    row += 1

    for item in reporte['desglose_egresos']:
        ws_resumen[f'A{row}'] = item['categoria']
        ws_resumen[f'B{row}'] = item['monto']
        ws_resumen[f'B{row}'].number_format = '#,##0.00'
        row += 1

    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 20

    # ============================================
    # HOJA 2: INGRESOS
    # ============================================
    ws_ingresos = wb.create_sheet("Ingresos")

    # Headers
    headers = ['Fecha', 'Descripcion', 'Monto', 'Categoria', 'Subcategoria',
               'Confianza', 'Persona_Nombre', 'Documento', 'Es_DEBIN', 'DEBIN_ID', 'Batch_ID']

    for col, header in enumerate(headers, start=1):
        cell = ws_ingresos.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Obtener movimientos INGRESOS
    movimientos_ing = db.query(Movimiento).filter(
        Movimiento.categoria == 'INGRESOS',
        extract('year', Movimiento.fecha) == int(year),
        extract('month', Movimiento.fecha) == int(month)
    ).order_by(Movimiento.fecha.desc()).all()

    row = 2
    for mov in movimientos_ing:
        ws_ingresos.cell(row=row, column=1, value=mov.fecha.strftime('%d/%m/%Y'))
        ws_ingresos.cell(row=row, column=2, value=mov.descripcion)
        ws_ingresos.cell(row=row, column=3, value=mov.monto).number_format = '#,##0.00'
        ws_ingresos.cell(row=row, column=4, value=mov.categoria or '')
        ws_ingresos.cell(row=row, column=5, value=mov.subcategoria or '')
        ws_ingresos.cell(row=row, column=6, value=mov.confianza_porcentaje or 0)
        ws_ingresos.cell(row=row, column=7, value=mov.persona_nombre or '')
        ws_ingresos.cell(row=row, column=8, value=mov.documento or '')
        ws_ingresos.cell(row=row, column=9, value='Si' if mov.es_debin else 'No')
        ws_ingresos.cell(row=row, column=10, value=mov.debin_id or '')
        ws_ingresos.cell(row=row, column=11, value=mov.batch_id or '')
        row += 1

    # Ajustar anchos
    ws_ingresos.column_dimensions['A'].width = 12
    ws_ingresos.column_dimensions['B'].width = 40
    ws_ingresos.column_dimensions['C'].width = 15

    # ============================================
    # HOJA 3: EGRESOS
    # ============================================
    ws_egresos = wb.create_sheet("Egresos")

    # Headers
    for col, header in enumerate(headers, start=1):
        cell = ws_egresos.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Obtener movimientos EGRESOS
    movimientos_egr = db.query(Movimiento).filter(
        Movimiento.categoria == 'EGRESOS',
        extract('year', Movimiento.fecha) == int(year),
        extract('month', Movimiento.fecha) == int(month)
    ).order_by(Movimiento.fecha.desc()).all()

    row = 2
    for mov in movimientos_egr:
        ws_egresos.cell(row=row, column=1, value=mov.fecha.strftime('%d/%m/%Y'))
        ws_egresos.cell(row=row, column=2, value=mov.descripcion)
        ws_egresos.cell(row=row, column=3, value=mov.monto).number_format = '#,##0.00'
        ws_egresos.cell(row=row, column=4, value=mov.categoria or '')
        ws_egresos.cell(row=row, column=5, value=mov.subcategoria or '')
        ws_egresos.cell(row=row, column=6, value=mov.confianza_porcentaje or 0)
        ws_egresos.cell(row=row, column=7, value=mov.persona_nombre or '')
        ws_egresos.cell(row=row, column=8, value=mov.documento or '')
        ws_egresos.cell(row=row, column=9, value='Si' if mov.es_debin else 'No')
        ws_egresos.cell(row=row, column=10, value=mov.debin_id or '')
        ws_egresos.cell(row=row, column=11, value=mov.batch_id or '')
        row += 1

    # Ajustar anchos
    ws_egresos.column_dimensions['A'].width = 12
    ws_egresos.column_dimensions['B'].width = 40
    ws_egresos.column_dimensions['C'].width = 15

    # ============================================
    # HOJA 4: TOP EGRESOS
    # ============================================
    ws_top = wb.create_sheet("Top Egresos")

    # Headers
    top_headers = ['Ranking', 'Fecha', 'Descripcion', 'Subcategoria', 'Monto', 'Batch_ID']
    for col, header in enumerate(top_headers, start=1):
        cell = ws_top.cell(row=1, column=col, value=header)
        cell.font = header_font

    # TOP 15 egresos por abs(monto)
    top_egresos = db.query(Movimiento).filter(
        Movimiento.categoria == 'EGRESOS',
        extract('year', Movimiento.fecha) == int(year),
        extract('month', Movimiento.fecha) == int(month)
    ).order_by(func.abs(Movimiento.monto).desc()).limit(15).all()

    row = 2
    for idx, mov in enumerate(top_egresos, start=1):
        ws_top.cell(row=row, column=1, value=idx)
        ws_top.cell(row=row, column=2, value=mov.fecha.strftime('%d/%m/%Y'))
        ws_top.cell(row=row, column=3, value=mov.descripcion)
        ws_top.cell(row=row, column=4, value=mov.subcategoria or '')
        ws_top.cell(row=row, column=5, value=mov.monto).number_format = '#,##0.00'
        ws_top.cell(row=row, column=6, value=mov.batch_id or '')
        row += 1

    # Ajustar anchos
    ws_top.column_dimensions['A'].width = 10
    ws_top.column_dimensions['B'].width = 12
    ws_top.column_dimensions['C'].width = 40
    ws_top.column_dimensions['D'].width = 20
    ws_top.column_dimensions['E'].width = 15

    # ============================================
    # HOJA 5: SIN CLASIFICAR
    # ============================================
    ws_sin_clas = wb.create_sheet("Sin Clasificar")

    # Headers
    sin_clas_headers = ['Fecha', 'Descripcion', 'Monto', 'Batch_ID']
    for col, header in enumerate(sin_clas_headers, start=1):
        cell = ws_sin_clas.cell(row=1, column=col, value=header)
        cell.font = header_font

    # Movimientos sin clasificar
    sin_clasificar = db.query(Movimiento).filter(
        extract('year', Movimiento.fecha) == int(year),
        extract('month', Movimiento.fecha) == int(month)
    ).filter(
        (Movimiento.categoria == None) |
        (Movimiento.categoria == '') |
        (Movimiento.categoria == 'SIN_CATEGORIA')
    ).order_by(Movimiento.fecha.desc()).all()

    row = 2
    for mov in sin_clasificar:
        ws_sin_clas.cell(row=row, column=1, value=mov.fecha.strftime('%d/%m/%Y'))
        ws_sin_clas.cell(row=row, column=2, value=mov.descripcion)
        ws_sin_clas.cell(row=row, column=3, value=mov.monto).number_format = '#,##0.00'
        ws_sin_clas.cell(row=row, column=4, value=mov.batch_id or '')
        row += 1

    # Ajustar anchos
    ws_sin_clas.column_dimensions['A'].width = 12
    ws_sin_clas.column_dimensions['B'].width = 40
    ws_sin_clas.column_dimensions['C'].width = 15

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def exportar_excel_ejecutivo(
    mes: str = Query(..., description="Mes en formato YYYY-MM"),
    db: Session = Depends(get_db)
):
    """
    Endpoint: GET /api/reportes/excel
    Exporta Excel Ejecutivo con 5 hojas completas
    """
    try:
        # Generar reporte (reutilizar lógica existente)
        from backend.core.reportes import generar_reporte_ejecutivo
        reporte = generar_reporte_ejecutivo(db, mes)

        # Generar Excel con 5 hojas
        excel_buffer = generar_excel_ejecutivo(reporte, mes, db)

        # Nombre del archivo
        filename = f"reporte_ejecutivo_{mes.replace('-', '_')}.xlsx"

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generando Excel Ejecutivo: {str(e)}"
        )
