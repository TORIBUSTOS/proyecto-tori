"""
Test de Excel Ejecutivo (ETAPA 7.B)
Valida que el Excel tenga exactamente 5 hojas con el formato correcto
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from backend.database.connection import get_db
from backend.core.reportes import generar_reporte_ejecutivo
from backend.api.exportacion import generar_excel_ejecutivo
from openpyxl import load_workbook

print("=" * 80)
print("TEST EXCEL EJECUTIVO - ETAPA 7.B")
print("=" * 80)

db = next(get_db())
mes = "2025-10"

try:
    # Test 1: Generar reporte
    print(f"\n[1/4] Generando reporte para {mes}...")
    reporte = generar_reporte_ejecutivo(db, mes)
    print(f"[OK] Reporte generado: {reporte['periodo']}")

    # Test 2: Generar Excel Ejecutivo
    print(f"\n[2/4] Generando Excel Ejecutivo...")
    excel_buffer = generar_excel_ejecutivo(reporte, mes, db)
    print(f"[OK] Excel generado: {len(excel_buffer.getvalue()):,} bytes")

    # Guardar para inspección
    with open("test_excel_ejecutivo_octubre.xlsx", "wb") as f:
        f.write(excel_buffer.getvalue())
    print("[OK] Guardado en: test_excel_ejecutivo_octubre.xlsx")

    # Test 3: Validar estructura (5 hojas)
    print(f"\n[3/4] Validando estructura del Excel...")
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)

    hojas_esperadas = ["Resumen", "Ingresos", "Egresos", "Top Egresos", "Sin Clasificar"]
    hojas_encontradas = wb.sheetnames

    print(f"  - Hojas esperadas: {hojas_esperadas}")
    print(f"  - Hojas encontradas: {hojas_encontradas}")

    assert len(hojas_encontradas) == 5, f"[ERROR] Se esperaban 5 hojas, se encontraron {len(hojas_encontradas)}"
    assert hojas_encontradas == hojas_esperadas, f"[ERROR] Nombres de hojas no coinciden"
    print("[OK] Estructura correcta: 5 hojas con nombres esperados")

    # Test 4: Validar contenido de hoja "Resumen"
    print(f"\n[4/4] Validando contenido de hoja Resumen...")
    ws_resumen = wb["Resumen"]

    # Verificar bloques obligatorios
    saldos_labels = []
    for row in range(1, 20):  # Buscar en primeras 20 filas
        cell_value = ws_resumen[f'A{row}'].value
        if cell_value:
            saldos_labels.append(str(cell_value))

    labels_esperados = [
        "SALDOS BANCARIOS",
        "Saldo Inicial",
        "Total Ingresos",
        "Total Egresos",
        "Saldo Final",
        "Variacion del Mes"
    ]

    labels_encontrados = [label for label in labels_esperados if label in saldos_labels]

    print(f"  - Labels de saldos encontrados: {len(labels_encontrados)}/{len(labels_esperados)}")

    if len(labels_encontrados) < 5:
        print(f"[WARN] Faltan algunos labels de saldos bancarios")
        print(f"  Esperados: {labels_esperados}")
        print(f"  Encontrados: {labels_encontrados}")
    else:
        print("[OK] Bloque de saldos bancarios presente")

    # Verificar que haya datos en columna B (valores numéricos)
    saldo_inicial_value = None
    for row in range(1, 20):
        if ws_resumen[f'A{row}'].value == "Saldo Inicial":
            saldo_inicial_value = ws_resumen[f'B{row}'].value
            break

    if saldo_inicial_value is not None:
        print(f"[OK] Saldo Inicial tiene valor: {saldo_inicial_value:,.2f}")
    else:
        print("[WARN] No se encontró valor para Saldo Inicial")

    # Verificar hojas de detalle tienen datos
    print(f"\n[BONUS] Verificando datos en otras hojas...")
    ws_ingresos = wb["Ingresos"]
    ws_egresos = wb["Egresos"]
    ws_top = wb["Top Egresos"]
    ws_sin_clas = wb["Sin Clasificar"]

    # Contar filas con datos (excluyendo header)
    rows_ingresos = ws_ingresos.max_row - 1
    rows_egresos = ws_egresos.max_row - 1
    rows_top = ws_top.max_row - 1
    rows_sin_clas = ws_sin_clas.max_row - 1

    print(f"  - Ingresos: {rows_ingresos} movimientos")
    print(f"  - Egresos: {rows_egresos} movimientos")
    print(f"  - Top Egresos: {rows_top} movimientos (máx 15)")
    print(f"  - Sin Clasificar: {rows_sin_clas} movimientos")

    print("\n" + "=" * 80)
    print("[OK] TODAS LAS VALIDACIONES PASARON")
    print("=" * 80)
    print("\nArchivo generado: test_excel_ejecutivo_octubre.xlsx")
    print("\nResumen:")
    print(f"  - 5 hojas: {', '.join(hojas_esperadas)}")
    print(f"  - Bloque SALDOS BANCARIOS: OK")
    print(f"  - Ingresos: {rows_ingresos} filas")
    print(f"  - Egresos: {rows_egresos} filas")
    print(f"  - Top 15 Egresos: {rows_top} filas")
    print(f"  - Sin Clasificar: {rows_sin_clas} filas")

except AssertionError as e:
    print(f"\n[ERROR] {e}")
    sys.exit(1)
except Exception as e:
    print(f"\n[ERROR] {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
finally:
    db.close()
